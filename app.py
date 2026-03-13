import streamlit as st
import openpyxl
import io

st.title("Transferência de Planilhas")
st.write("Suba o Excel A e baixe o Excel B já preenchido")

arquivo = st.file_uploader("📎 Suba o Excel A aqui", type=["xlsx"])

if arquivo:
# Lê o Excel A
wb_origem = openpyxl.load_workbook(arquivo)
aba_origem = wb_origem.active

# Cria o Excel B
wb_destino = openpyxl.Workbook()
aba_destino = wb_destino.active

# Copia todas as linhas de A para B
for linha in aba_origem.iter_rows(values_only=True):
aba_destino.append(linha)

# Gera o arquivo
output = io.BytesIO()
wb_destino.save(output)
output.seek(0)

st.success("✅ Dados transferidos com sucesso!")

st.download_button(
label="⬇️ Baixar Excel B preenchido",
data=output,
file_name="planilha_destino.xlsx",
mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
